"""Pydantic-модели для операций и планов обработки Excel-файлов."""

from __future__ import annotations

from abc import ABC, abstractmethod
from typing import Literal, Union, Annotated
from pydantic import BaseModel, Field


class ColumnReference(BaseModel):
    """Ссылка на столбец Excel."""

    name: str = Field(description="Имя столбца, как оно указано в заголовке Excel")
    excel_column: str = Field(description="Буква столбца в Excel (например, 'B')")


# ---------------------------------------------------------------------------
# Базовая операция
# ---------------------------------------------------------------------------

class BaseOperation(BaseModel, ABC):
    """Базовый абстрактный класс для всех операций."""

    DESCRIPTION: str = Field(default="", exclude=True)
    args: dict = Field(default_factory=dict)

    @classmethod
    @abstractmethod
    def parse_args(cls, json_data: dict) -> BaseOperation:
        """Создаёт экземпляр из JSON-словаря."""
        ...

    @abstractmethod
    def execute(self, input_file: str, output_file: str) -> None:
        """Выполняет операцию."""
        ...


# ---------------------------------------------------------------------------
# Конкретные операции
# ---------------------------------------------------------------------------

class BarChartOperation(BaseOperation):
    """Операция создания столбчатой диаграммы."""

    DESCRIPTION: str = Field(default="Создаёт столбчатую диаграмму по указанным столбцам данных", exclude=True)

    class Args(BaseModel):
        """Аргументы для создания столбчатой диаграммы."""

        title: str = Field(description="Заголовок диаграммы")
        data_columns: list[ColumnReference] = Field(description="Столбцы со значениями (ось Y)")
        label_column: ColumnReference = Field(description="Столбец с подписями (ось X)")
        output_sheet: str = Field(default="Диаграмма", description="Имя листа для диаграммы")

    operation: Literal["СТОЛБЧАТАЯ_ДИАГРАММА"]
    args: Args

    @classmethod
    def parse_args(cls, json_data: dict) -> BarChartOperation:
        """Создаёт экземпляр из JSON-словаря."""
        args = cls.Args(**json_data.get("args", json_data))
        return cls(operation=json_data["operation"], args=args)

    def execute(self, input_file: str, output_file: str) -> None:
        """Создаёт столбчатую диаграмму в Excel-файле."""
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import column_index_from_string

        a = self.args
        wb = load_workbook(input_file)
        ws = wb.active

        if a.output_sheet in wb.sheetnames:
            del wb[a.output_sheet]
        chart_ws = wb.create_sheet(a.output_sheet)

        cats_ref = Reference(
            ws,
            min_col=column_index_from_string(a.label_column.excel_column),
            min_row=2, max_row=ws.max_row,
        )

        chart = BarChart()
        chart.title = a.title
        chart.style = 10
        chart.grouping = "clustered"

        num_rows = ws.max_row - 1
        num_cols = len(a.data_columns)
        chart.width = max(18, min(40, 10 + num_rows * 1.2))
        chart.height = max(10, min(20, 8 + num_cols * 1.5))

        chart.x_axis.title = a.label_column.name
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.delete = False

        data_names = []
        for col in a.data_columns:
            data_ref = Reference(
                ws,
                min_col=column_index_from_string(col.excel_column),
                min_row=1, max_row=ws.max_row,
            )
            chart.add_data(data_ref, titles_from_data=True)
            data_names.append(col.name)

        chart.y_axis.title = ", ".join(data_names) if len(data_names) != 1 else data_names[0]
        chart.y_axis.numFmt = '#,##0'
        chart.y_axis.tickLblPos = "low"
        chart.y_axis.delete = False

        chart.set_categories(cats_ref)
        chart_ws.add_chart(chart, "A1")
        wb.save(output_file)

        cols = ", ".join(c.name for c in a.data_columns)
        print(f"[EXECUTOR] Гистограмма '{a.title}' [{cols}] → лист '{a.output_sheet}'")


class PieChartOperation(BaseOperation):
    """Операция создания круговой диаграммы."""

    DESCRIPTION: str = Field(default="Создаёт круговую диаграмму (секторную) по указанным столбцам данных", exclude=True)

    class Args(BaseModel):
        """Аргументы для создания круговой диаграммы."""

        title: str = Field(description="Заголовок диаграммы")
        data_column: ColumnReference = Field(description="Столбец со значениями")
        label_column: ColumnReference = Field(description="Столбец с подписями (секторов)")
        output_sheet: str = Field(default="Диаграмма", description="Имя листа для диаграммы")

    operation: Literal["КРУГОВАЯ_ДИАГРАММА"]
    args: Args

    @classmethod
    def parse_args(cls, json_data: dict) -> PieChartOperation:
        """Создаёт экземпляр из JSON-словаря."""
        args = cls.Args(**json_data.get("args", json_data))
        return cls(operation=json_data["operation"], args=args)

    def execute(self, input_file: str, output_file: str) -> None:
        """Создаёт круговую диаграмму в Excel-файле."""
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart, Reference
        from openpyxl.utils import column_index_from_string

        a = self.args
        wb = load_workbook(input_file)
        ws = wb.active

        if a.output_sheet in wb.sheetnames:
            del wb[a.output_sheet]
        chart_ws = wb.create_sheet(a.output_sheet)

        data_ref = Reference(
            ws,
            min_col=column_index_from_string(a.data_column.excel_column),
            min_row=1, max_row=ws.max_row,
        )
        cats_ref = Reference(
            ws,
            min_col=column_index_from_string(a.label_column.excel_column),
            min_row=2, max_row=ws.max_row,
        )

        chart = PieChart()
        chart.title = a.title
        chart.style = 10
        chart.width = 18
        chart.height = 14
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart_ws.add_chart(chart, "A1")
        wb.save(output_file)
        print(f"[EXECUTOR] Круговая диаграмма '{a.title}' → лист '{a.output_sheet}'")


class ArithmeticOperation(BaseOperation):
    """Операция арифметических вычислений."""

    DESCRIPTION: str = Field(default="Выполняет арифметическую операцию (сложение, вычитание, умножение, деление) над двумя столбцами и записывает результат в новый столбец", exclude=True)

    class Args(BaseModel):
        """Аргументы для арифметической операции."""

        operation_type: Literal["СЛОЖЕНИЕ", "ВЫЧИТАНИЕ", "УМНОЖЕНИЕ", "ДЕЛЕНИЕ"]
        operand1: ColumnReference = Field(description="Первый операнд (столбец)")
        operand2: ColumnReference = Field(description="Второй операнд (столбец)")
        result_column_name: str = Field(description="Имя нового столбца с результатом")
        result_column_letter: str = Field(description="Буква столбца для результата (например, 'E')")
        output_sheet: str = Field(default="", description="Лист для записи. Пусто = активный лист")

    operation: Literal["АРИФМЕТИКА"]
    args: Args

    @classmethod
    def parse_args(cls, json_data: dict) -> ArithmeticOperation:
        """Создаёт экземпляр из JSON-словаря."""
        args = cls.Args(**json_data.get("args", json_data))
        return cls(operation=json_data["operation"], args=args)

    def execute(self, input_file: str, output_file: str) -> None:
        """Выполняет арифметическую операцию и записывает результат в столбец."""
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string

        a = self.args
        wb = load_workbook(input_file)
        ws = wb.active if not a.output_sheet else (
            wb[a.output_sheet] if a.output_sheet in wb.sheetnames else wb.active
        )

        result_col_idx = column_index_from_string(a.result_column_letter)
        col1_idx = column_index_from_string(a.operand1.excel_column)
        col2_idx = column_index_from_string(a.operand2.excel_column)

        ws.cell(row=1, column=result_col_idx, value=a.result_column_name)

        for row in range(2, ws.max_row + 1):
            v1 = ws.cell(row=row, column=col1_idx).value or 0
            v2 = ws.cell(row=row, column=col2_idx).value or 0

            if a.operation_type == "СЛОЖЕНИЕ":
                result = v1 + v2
            elif a.operation_type == "ВЫЧИТАНИЕ":
                result = v1 - v2
            elif a.operation_type == "УМНОЖЕНИЕ":
                result = v1 * v2
            elif a.operation_type == "ДЕЛЕНИЕ":
                result = v1 / v2 if v2 != 0 else "#DIV/0!"
            else:
                result = "#UNKNOWN!"

            ws.cell(row=row, column=result_col_idx, value=result)

        wb.save(output_file)
        print(f"[EXECUTOR] Арифметика: {a.operand1.name} {a.operation_type} {a.operand2.name} → '{a.result_column_name}' [столбец {a.result_column_letter}]")


class ConditionalFormatOperation(BaseOperation):
    """Операция условного форматирования ячеек."""

    DESCRIPTION: str = Field(default="Форматирует ячейки столбца по одному из условий: БОЛЬШЕ, МЕНЬШЕ, РАВНО, НЕ_РАВНО, БОЛЬШЕ_ИЛИ_РАВНО, МЕНЬШЕ_ИЛИ_РАВНО", exclude=True)

    class Args(BaseModel):
        """Аргументы для условного форматирования."""

        target_column: ColumnReference = Field(description="Столбец, к которому применяется форматирование")
        operator: Literal["БОЛЬШЕ", "МЕНЬШЕ", "РАВНО", "НЕ_РАВНО", "БОЛЬШЕ_ИЛИ_РАВНО", "МЕНЬШЕ_ИЛИ_РАВНО"]
        value: Union[float, str] = Field(description="Значение для сравнения (число или строка)")
        fill_color: str = Field(default="FF0000", description="Цвет фона в hex без # (например, C6EFCE = зелёный)")
        font_color: str = Field(default="000000", description="Цвет текста в hex без #")

    operation: Literal["УСЛОВНОЕ_ФОРМАТИРОВАНИЕ"]
    args: Args

    @classmethod
    def parse_args(cls, json_data: dict) -> ConditionalFormatOperation:
        args = cls.Args(**json_data.get("args", json_data))
        return cls(operation=json_data["operation"], args=args)

    def execute(self, input_file: str, output_file: str) -> None:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill, Font

        a = self.args
        wb = load_workbook(input_file)
        ws = wb.active

        col_letter = a.target_column.excel_column
        cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        OPERATOR_MAP = {
            "БОЛЬШЕ": "greaterThan",
            "МЕНЬШЕ": "lessThan",
            "РАВНО": "equal",
            "НЕ_РАВНО": "notEqual",
            "БОЛЬШЕ_ИЛИ_РАВНО": "greaterThanOrEqual",
            "МЕНЬШЕ_ИЛИ_РАВНО": "lessThanOrEqual",
        }

        rule = CellIsRule(
            operator=OPERATOR_MAP[a.operator],
            formula=[str(a.value)],
            fill=PatternFill(start_color=a.fill_color, end_color=a.fill_color, fill_type="solid"),
            font=Font(color=a.font_color),
        )
        ws.conditional_formatting.add(cell_range, rule)

        wb.save(output_file)
        print(f"[EXECUTOR] Условное форматирование: {a.target_column.name} {a.operator} {a.value} → {col_letter}")


class UndefinedOperation(BaseOperation):
    """Операция-заглушка для неопределённых задач."""

    DESCRIPTION: str = Field(default="Операция не может быть выполнена (заглушка)", exclude=True)

    class Args(BaseModel):
        """Аргументы для неопределённой операции."""

        reason: str

    operation: Literal["НЕ_ОПРЕДЕЛЕНО"]
    args: Args

    @classmethod
    def parse_args(cls, json_data: dict) -> UndefinedOperation:
        """Создаёт экземпляр из JSON-словаря."""
        args = cls.Args(**json_data.get("args", json_data))
        return cls(operation=json_data["operation"], args=args)

    def execute(self, input_file: str, output_file: str) -> None:
        """Выводит сообщение о пропуске операции."""
        print(f"[EXECUTOR] Пропуск: {self.args.reason}")


# ---------------------------------------------------------------------------
# Единый реестр операций
# ---------------------------------------------------------------------------

OPERATION_CLASSES: list[type[BaseOperation]] = [
    BarChartOperation,
    PieChartOperation,
    ArithmeticOperation,
    ConditionalFormatOperation,
    UndefinedOperation,
]

OPERATION_MAP: dict[str, type[BaseOperation]] = {cls.__name__: cls for cls in OPERATION_CLASSES}


# ---------------------------------------------------------------------------
# Discriminated union + Plan
# ---------------------------------------------------------------------------

ExcelOperation = Annotated[
    Union[tuple(OPERATION_CLASSES)],
    Field(discriminator="operation"),
]


class TaskStep(BaseModel):
    """Шаг декомпозиции задачи."""

    step: int = Field(description="Номер действия")
    action: str = Field(description="Что нужно сделать")


class TemplateTaskDecomposition(BaseModel):
    """Модель для декомпозиции задачи на отдельные действия."""

    steps: list[TaskStep]


class OperationSelection(BaseModel):
    """Выбор одной операции для одной подзадачи."""

    operation: str = Field(description="Имя операции из доступного списка")


class ExcelPlan(BaseModel):
    """Модель плана обработки Excel-файла."""

    operations: list[ExcelOperation]
    summary: str = Field(description="Краткое описание того, что будет сделано")



