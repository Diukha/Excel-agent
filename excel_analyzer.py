import openpyxl
import json
from openpyxl.utils import get_column_letter

def analyze_excel(file_path, max_sample_rows=3, max_unique_values=5):
    """
    Создает максимально компактное JSON-представление структуры Excel файла.
    Идеально для передачи в контекст ИИ-агента.
    """
    analysis_results = {"sheets": {}}
    
    try:
        # data_only=True возвращает значения формул, а не сами формулы (экономит токены и дает реальные данные)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 1. Находим реальные границы данных (bounding box), игнорируя пустые края
            min_row, max_row, min_col, max_col = None, None, None, None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        r, c = cell.row, cell.column
                        if min_row is None or r < min_row: min_row = r
                        if max_row is None or r > max_row: max_row = r
                        if min_col is None or c < min_col: min_col = c
                        if max_col is None or c > max_col: max_col = c
                        
            if min_row is None:
                analysis_results["sheets"][sheet_name] = "Лист пуст"
                continue
            
            # 2. Извлекаем заголовки (первая строка найденного блока)
            headers = []
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=min_row, column=c)
                val = cell.value
                col_name = str(val).strip() if val else f"Column_{c}"
                headers.append({
                    "excel_column": get_column_letter(c),
                    "name": col_name
                })
            
            # 3. Примеры данных (СТРОГО ограничено max_sample_rows, по умолчанию 3)
            sample_data = []
            sample_end_row = min(min_row + max_sample_rows, max_row)
            for row in ws.iter_rows(min_row=min_row + 1, max_row=sample_end_row, min_col=min_col, max_col=max_col):
                row_dict = {"_excel_row": row[0].row} # Реальный номер строки в Excel
                for i, cell in enumerate(row):
                    col_name = headers[i]["name"]
                    val = cell.value
                    # Преобразуем в строку только если это не None, чтобы JSON оставался чистым
                    row_dict[col_name] = str(val) if val is not None else None
                sample_data.append(row_dict)
            
            # 4. Структура колонок (только статистика, БЕЗ выгрузки всех данных)
            columns_info = []
            for i, header in enumerate(headers):
                col_letter = header["excel_column"]
                col_name = header["name"]
                
                non_null_count = 0
                null_count = 0
                unique_vals = set()
                
                # Сканируем только текущий столбец в пределах блока данных
                for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col + i, max_col=min_col + i):
                    cell = row[0]
                    if cell.value is not None and str(cell.value).strip() != "":
                        non_null_count += 1
                        # Собираем уникальные значения, но ограничиваем размер множества, чтобы не тратить память
                        if len(unique_vals) < 5: 
                            unique_vals.add(str(cell.value))
                    else:
                        null_count += 1
                        
                col_info = {
                    "name": col_name,
                    "excel_column": col_letter,
                    "non_null_count": non_null_count,
                    "null_count": null_count,
                }
                
                # Добавляем уникальные значения ТОЛЬКО если их действительно мало (это категории/статусы)
                unique_list = list(unique_vals)
                if len(unique_list) <= max_unique_values:
                    col_info["unique_values"] = unique_list
                
                columns_info.append(col_info)
                
            # 5. Формируем итоговый компактный словарь для листа
            analysis_results["sheets"][sheet_name] = {
                "data_block": {
                    "start_row": min_row,
                    "end_row": max_row,
                    "start_col": get_column_letter(min_col),
                    "end_col": get_column_letter(max_col)
                },
                "total_rows_in_block": max_row - min_row,
                "columns": columns_info,
                "sample_data": sample_data # Всего 3 строки, как и просили
            }
            
        return analysis_results
        
    except Exception as e:
        return {"error": f"Не удалось прочитать файл: {str(e)}"}

# Пример использования:
if __name__ == "__main__":
    report = analyze_excel_structure("data.xlsx")
    # Выводим результат, убеждаясь, что он компактен
    print(json.dumps(report, indent=2, ensure_ascii=False))